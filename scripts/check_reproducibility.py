#!/usr/bin/env python3
"""Run consolidation twice and require stable reports/exports and workbook semantics."""

from __future__ import annotations

import hashlib
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PYTHON = sys.executable
DETERMINISTIC_ROOTS = [ROOT / "docs", ROOT / "outputs" / "machine-readable", ROOT / "outputs" / "reports", ROOT / "outputs" / "MASTER_DATA_REFERENCE.md", ROOT / "working" / "source_hashes.json"]
WORKBOOK = ROOT / "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx"


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def deterministic_hashes() -> dict[str, str]:
    files: list[Path] = []
    for root in DETERMINISTIC_ROOTS:
        if root.is_file(): files.append(root)
        elif root.exists(): files.extend(p for p in root.rglob("*") if p.is_file())
    return {str(p.relative_to(ROOT)): digest(p) for p in sorted(set(files))}


def workbook_signature() -> str:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    payload = []
    for ws in wb.worksheets:
        payload.append({
            "name": ws.title, "state": ws.sheet_state, "dimensions": [ws.max_row, ws.max_column],
            "values": [[cell.value for cell in row] for row in ws.iter_rows()],
            "merges": sorted(str(x) for x in ws.merged_cells.ranges), "freeze": str(ws.freeze_panes or ""),
            "filter": str(ws.auto_filter.ref or ""), "tables": sorted(ws.tables.keys()),
        })
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def run() -> tuple[dict[str, str], str]:
    subprocess.run([PYTHON, "scripts/consolidate.py"], cwd=ROOT, check=True)
    return deterministic_hashes(), workbook_signature()


def main() -> None:
    first_files, first_workbook = run()
    second_files, second_workbook = run()
    result = {
        "stable_deterministic_files": first_files == second_files,
        "stable_workbook_semantics": first_workbook == second_workbook,
        "deterministic_file_count": len(first_files),
        "workbook_signature": second_workbook,
        "changed_files": sorted(k for k in set(first_files) | set(second_files) if first_files.get(k) != second_files.get(k)),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if not all([result["stable_deterministic_files"], result["stable_workbook_semantics"]]):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
