#!/usr/bin/env python3
"""Validate required deliverables, syntax, workbook structure, and source hashes."""

from __future__ import annotations

import csv
import hashlib
import json
import subprocess
import zipfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
REQUIRED = [
    "AGENTS.md", "PROJECT_STATE.md", "docs/REPOSITORY_INVENTORY.md",
    "docs/prompts/BANK_VALIDATION_TASK.md", "docs/prompts/NEW_SOURCE_INGESTION_TASK.md",
    "outputs/MASTER_DATA_REFERENCE.md",
    "outputs/reports/MISSING_INFORMATION.md", "outputs/reports/CONFLICTS_AND_DECISIONS.md",
    "outputs/reports/CHANGELOG.md", "outputs/reports/WORKBOOK_AUDIT.md",
    "outputs/reports/COLLECTION_STATUS.md", "outputs/reports/FINAL_VALIDATION_PLAN.md",
    "outputs/machine-readable/cards.csv", "outputs/machine-readable/cards.json",
    "outputs/machine-readable/sources.csv", "outputs/machine-readable/conflicts.csv",
    "outputs/machine-readable/missing_fields.csv", "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx",
]


def digest(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""): h.update(block)
    return h.hexdigest()


def main():
    checks = []
    for rel in REQUIRED:
        p = ROOT / rel
        checks.append({"check": f"exists:{rel}", "passed": p.exists() and p.stat().st_size > 0})
    cards = json.loads((ROOT / "outputs/machine-readable/cards.json").read_text(encoding="utf-8"))
    checks.append({"check": "cards.json valid nonempty array", "passed": isinstance(cards, list) and len(cards) > 0, "count": len(cards)})
    for name in ["cards.csv", "sources.csv", "conflicts.csv", "missing_fields.csv"]:
        with (ROOT / "outputs/machine-readable" / name).open(encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))
        checks.append({"check": f"CSV valid:{name}", "passed": len(rows) > 0, "count": len(rows)})
    original = load_workbook(ROOT / "Credit Cards Terms and Conditions/01. saudi-credit-cards-unified V3.xlsx", data_only=False)
    output = load_workbook(ROOT / "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx", data_only=False)
    checks.append({"check": "all original sheet names retained", "passed": set(original.sheetnames) <= set(output.sheetnames)})
    checks.append({"check": "seven additive sheets present", "passed": len(output.sheetnames) == len(original.sheetnames) + 7,
                   "original": len(original.sheetnames), "output": len(output.sheetnames)})
    checks.append({"check": "ANB Phase 2 validation sheet present", "passed": "ANB Validation 2026-07-30" in output.sheetnames})
    checks.append({"check": "original card row count not decreased", "passed": output["دليل البطاقات"].max_row >= original["دليل البطاقات"].max_row,
                   "original": original["دليل البطاقات"].max_row, "output": output["دليل البطاقات"].max_row})
    old_formulas = sum(c.data_type == "f" for ws in original.worksheets for row in ws.iter_rows() for c in row)
    new_formulas_original_sheets = sum(c.data_type == "f" for name in original.sheetnames for row in output[name].iter_rows() for c in row)
    checks.append({"check": "original formula count preserved", "passed": old_formulas == new_formulas_original_sheets,
                   "original": old_formulas, "output_original_sheets": new_formulas_original_sheets})
    snapshot = json.loads((ROOT / "working/source_hashes.json").read_text(encoding="utf-8"))
    source_checks = []
    for row in snapshot:
        if row["path"].startswith("Credit Cards Terms and Conditions/"):
            p = ROOT / row["path"]
            source_checks.append(p.exists() and digest(p) == row["sha256"])
    checks.append({"check": "source files unchanged since inventory", "passed": all(source_checks), "count": len(source_checks)})
    raw_diff = subprocess.run(
        ["git", "status", "--porcelain", "--", "Credit Cards Terms and Conditions"],
        cwd=ROOT, capture_output=True, text=True, check=True,
    ).stdout.strip()
    checks.append({"check": "raw source folder has no Git changes", "passed": raw_diff == "", "details": raw_diff})
    raw_names = {p.name for p in (ROOT / "Credit Cards Terms and Conditions").iterdir() if p.is_file()}
    generated_names = {"saudi-credit-cards-unified-consolidated.xlsx", "MASTER_DATA_REFERENCE.md", "MISSING_INFORMATION.md",
                       "CONFLICTS_AND_DECISIONS.md", "CHANGELOG.md", "WORKBOOK_AUDIT.md", "COLLECTION_STATUS.md",
                       "FINAL_VALIDATION_PLAN.md", "cards.csv", "cards.json", "sources.csv", "conflicts.csv", "missing_fields.csv"}
    checks.append({"check": "generated deliverables are outside raw source folder", "passed": not bool(raw_names & generated_names),
                   "unexpected": sorted(raw_names & generated_names)})
    workbook_path = ROOT / "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx"
    with zipfile.ZipFile(workbook_path) as archive:
        bad_member = archive.testzip()
    checks.append({"check": "consolidated XLSX ZIP integrity", "passed": bad_member is None, "bad_member": bad_member})
    markdown_paths = [ROOT / rel for rel in REQUIRED if rel.endswith(".md")]
    checks.append({"check": "Markdown outputs decode as UTF-8", "passed": all(p.read_text(encoding="utf-8") is not None for p in markdown_paths),
                   "count": len(markdown_paths)})
    state_text = (ROOT / "PROJECT_STATE.md").read_text(encoding="utf-8")
    state_keys = ["repository:", "latest_origin_main_commit:", "recommended_next_bank:",
                  "open_pull_requests:", "exact_next_recommended_action:", "last_updated_commit:"]
    checks.append({"check": "PROJECT_STATE required keys present",
                   "passed": all(key in state_text for key in state_keys), "keys": state_keys})
    result = {"passed": all(c["passed"] for c in checks), "checks": checks}
    (ROOT / "outputs/reports/validation_results.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if not result["passed"]: raise SystemExit(1)


if __name__ == "__main__":
    main()
