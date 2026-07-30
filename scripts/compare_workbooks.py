#!/usr/bin/env python3
"""Compare protected original sheets with the consolidated workbook."""

from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ORIGINAL = ROOT / "Credit Cards Terms and Conditions/01. saudi-credit-cards-unified V3.xlsx"
OUTPUT = ROOT / "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx"


def sheet_signature(ws):
    cells = []
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None or c.has_style or c.comment or c.hyperlink:
                cells.append((c.coordinate, c.value, c.data_type, c.number_format, c.style_id,
                              c.comment.text if c.comment else None, c.hyperlink.target if c.hyperlink else None))
    return {
        "dimensions": (ws.max_row, ws.max_column), "cells": cells,
        "merged": sorted(str(x) for x in ws.merged_cells.ranges),
        "hidden_rows": sorted(k for k, v in ws.row_dimensions.items() if v.hidden),
        "hidden_columns": sorted(k for k, v in ws.column_dimensions.items() if v.hidden),
        "tables": sorted(ws.tables.keys()), "freeze": str(ws.freeze_panes or ""),
        "auto_filter": str(ws.auto_filter.ref or ""), "state": ws.sheet_state,
    }


def main():
    a = load_workbook(ORIGINAL, data_only=False)
    b = load_workbook(OUTPUT, data_only=False)
    results = []
    for name in a.sheetnames:
        same = sheet_signature(a[name]) == sheet_signature(b[name])
        results.append({"sheet": name, "preserved_semantically": same,
                        "original_rows": a[name].max_row, "output_rows": b[name].max_row,
                        "original_columns": a[name].max_column, "output_columns": b[name].max_column})
    report = {"original_sheets": a.sheetnames, "output_sheets": b.sheetnames,
              "original_sheet_deletions": len(set(a.sheetnames) - set(b.sheetnames)),
              "all_original_sheets_preserved": all(r["preserved_semantically"] for r in results), "sheets": results}
    path = ROOT / "outputs/reports/workbook_comparison.json"
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not report["all_original_sheets_preserved"] or report["original_sheet_deletions"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
