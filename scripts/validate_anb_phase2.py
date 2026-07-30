#!/usr/bin/env python3
"""Add the 2026-07-30 ANB official-source validation register.

This script only edits the generated consolidated workbook. It never reads from or
writes to the raw-source folder. Existing worksheets and rows are retained. The
validation sheet is recreated on rerun so the operation is deterministic at the
data level.
"""

from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs/excel/saudi-credit-cards-unified-consolidated.xlsx"
SHEET = "ANB Validation 2026-07-30"
VALIDATED = "2026-07-30"

TARIFF = "https://anb.com.sa/documents/55607/0/Tariff%2BLeaflet-EA%2BFA%2BReport%2BEN4.pdf/62b1f761-06a7-4527-1d7e-c36a9b28e313?t=1770817099979"
CREDIT_TC = "https://anb.com.sa/documents/55607/0/Credit%2BCard%2BTerms%2Band%2Bconditions%2B-%2BEN-%2BUpdated%2Bcopy.pdf/db369e63-2280-b727-db68-e678200cdb67?t=1736932162246"
REWARDS = "https://anb.com.sa/documents/55607/0/anb%2BRewards%2BProgram%2BT%26C%2Bsep.pdf/26d519c7-45b8-ef4c-4a74-c5ba4108fd4d?t=1758197765692"
LOYALTY = "https://anb.com.sa/documents/55607/0/Loyalty_Changes%2BMarketing%2B-%2BEng.pdf/1928e75b-fdcc-4228-205c-59b8979084ce?t=1751795153178"
ALFURSAN_TC = "https://anb.com.sa/documents/55607/0/Al%2BFursan%2BT%26C_EN_V2.pdf/6fa27f68-aede-af91-9563-ed665bcd0e77?t=1753109617919"
MADA_TC = "https://anb.com.sa/documents/55607/0/mada%2Bcard%2BEnglish%2BTC%2B26%2BJan.pdf/36b56695-6c3b-5988-fe2c-87959cac6afd?t=1770040280085"
CORP_TC = "https://anb.com.sa/documents/55607/0/ANB_Corporate%2BCredit%2BCard_T%26C_EN.pdf/8ae3ab8f-b1fd-a0ec-94ba-99ad7dc073fb?t=1753021943607"
CURRENCIES_TC = "https://anb.com.sa/documents/55607/0/ANB_Currencies%2BCard_T%26C_EN.pdf/b5e7dc75-85dd-5559-e9b0-bcd3ab51b63d?t=1753347817491"
LOW_LIMIT = "https://anb.com.sa/documents/55607/0/Low%2BLimit%2B_EN_v3.pdf/c784b40b-a05d-fcad-d82f-1c952aba04eb?t=1770200088929"
DIGITAL = "https://anb.com.sa/documents/55607/0/Digital%2BCashback_EN_V3.pdf/ed1e6aa3-dc59-2b49-53cd-a38d5cbd503e?t=1770200020066"


HEADERS = [
    "Card ID", "Official product name", "Network", "Product type",
    "Official-site status", "Annual fee / pricing evidence", "Reward evidence",
    "Validation outcome", "Open conflict or missing evidence", "Action taken",
    "Validated on", "Product page", "Pricing guide", "Applicable T&C",
    "Reward source",
]


def row(card_id, name, network, product_type, status, fee, reward, outcome,
        gap, action, page, terms=CREDIT_TC, reward_source=LOYALTY):
    return [card_id, name, network, product_type, status, fee, reward, outcome,
            gap, action, VALIDATED, page, TARIFF, terms, reward_source]


ROWS = [
    row("ANB-01", "anb Visa Infinite Privilege Credit Card", "Visa", "Credit", "Current product confirmed", "SAR 3,000 in 18-Feb-2026 tariff; product pages show 2.75% international while tariff says 2%", "1.8 points/SAR local; 1.8 vs 2.2 international in two official loyalty files", "Conflict pending", "International fee and international earn rate conflict", "Preserved both official values; no master overwrite", "https://anb.com.sa/web/anb/airport-lounges"),
    row("ANB-02", "Mastercard World Elite Exclusive Credit Card", "Mastercard", "Credit", "Current product confirmed by current offers", "Tariff prices 'MasterCard World Elite' at SAR 3,000; Exclusive attribution not explicit", "1.8 local; 1.8 vs 2.2 international", "Missing evidence", "Dedicated product/pricing disclosure for Exclusive is not published", "No identity merge or fee overwrite", "https://anb.com.sa/web/anb/zaps-offer-2"),
    row("ANB-03", "Mastercard World Elite Select Credit Card", "Mastercard", "Credit", "Presence confirmed only by current airport-transport page", "No product-specific fee; tariff only states World Elite SAR 3,000", "No Select-specific rate published", "Missing evidence", "Dedicated current product page, fee disclosure, T&C applicability and reward rate are missing", "No identity merge or inferred values", "https://anb.com.sa/web/anb/zaps-airport-transportation"),
    row("ANB-04", "anb Visa Infinite Credit Card", "Visa", "Credit", "Current product confirmed", "SAR 850 tariff/page; page 2.75% international vs tariff 2%", "1.5 local; 2.0 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/web/anb/visa-infinite"),
    row("ANB-05", "Mastercard World Credit Card", "Mastercard", "Credit", "Current product confirmed", "SAR 700 tariff/page; page 2.75% international vs tariff 2%", "1.2 local; 1.8 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/en/web/anb/mastercard-world-card"),
    row("ANB-06", "anb Visa Signature Credit Card", "Visa", "Credit", "Current product confirmed", "SAR 450 tariff; product page previously captured at 2.75% international vs tariff 2%", "1.2 local; 1.8 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/ar/web/anb/visa-signature"),
    row("ANB-07", "anb Visa Platinum Credit Card", "Visa", "Credit", "Current product confirmed", "SAR 300 tariff; product page previously captured at 2.75% international vs tariff 2%", "0.8 local; 1.5 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/web/anb/visa-platinum"),
    row("ANB-08", "anb Mastercard Platinum Credit Card", "Mastercard", "Credit", "Current product confirmed", "SAR 350 tariff; product page previously captured at 2.75% international vs tariff 2%", "0.8 local; 1.5 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/web/anb/mastercard-platinum"),
    row("ANB-09", "anb Mastercard Titanium Credit Card", "Mastercard", "Credit", "Current product confirmed", "SAR 200 tariff; product page previously captured at 2.75% international vs tariff 2%", "0.5 local; 0.8 international", "Conflict pending", "International transaction fee conflict", "Preserved both official values", "https://anb.com.sa/web/anb/mastercard-titanium"),
    row("ANB-10", "anb Visa Classic Credit Card", "Visa", "Credit", "Listed in current tariff and reward schedule; dedicated product page not located", "SAR 200 tariff", "0.5 local; 0.8 international", "Missing evidence", "Dedicated current product page/current availability statement is missing", "Kept as current-uncertain; no master overwrite", "https://anb.com.sa/web/anb/credit-cards"),
    row("ANB-11", "anb AlFursan Visa Infinite", "Visa", "Co-brand credit", "Current product confirmed", "SAR 2,000 tariff", "SAR 3 local/SAR 2 international per mile, plus documented milestone bonuses", "Validated with conflict backlog", "Product status/fee/reward supported; generic 2% vs page-level 2.75% fee conflict remains", "No destructive master edit", "https://anb.com.sa/web/anb/alfursan-infinite", ALFURSAN_TC, ALFURSAN_TC),
    row("ANB-12", "anb AlFursan Visa Signature", "Visa", "Co-brand credit", "Current family product confirmed", "SAR 1,000 tariff", "SAR 4 local/SAR 2.5 international per mile", "Validated with missing detail", "Dedicated current Signature page and exact active milestone-bonus wording not located", "Preserved existing provisional bonus detail", "https://anb.com.sa/web/anb/anb-alfursan", ALFURSAN_TC, ALFURSAN_TC),
    row("ANB-13", "Cashback Digital Card / E-Shopping Card", "Network not stated in official text", "Prepaid/digital", "Current product confirmed", "SAR 100; 2% international; SAR 15 replacement", "0.5% local; 1% international instant cashback", "Validated", "Network logo/text is not stated in accessible official text", "Added validated register only", "https://anb.com.sa/web/anb/e-shopping-credit-card", DIGITAL, DIGITAL),
    row("ANB-14", "Platinum Low-Limit Credit Card", "Visa", "Low-limit", "Current fee disclosure confirmed", "SAR 150 primary and supplementary; 2% international", "No reward rate published", "Missing evidence", "Dedicated product page and reward/non-reward statement missing", "No inferred reward value", "https://anb.com.sa/web/anb/visa-low-limit", LOW_LIMIT, LOW_LIMIT),
    row("ANB-15", "anb Low-Limit Credit Card", "Visa", "Low-limit", "Current product confirmed", "SAR 100 primary and supplementary", "No reward rate published", "Missing evidence", "Official disclosure has blank/non-applicable fee cells but no explicit reward statement", "No inferred reward value", "https://anb.com.sa/web/anb/visa-low-limit", LOW_LIMIT, LOW_LIMIT),
    row("ANB-16", "anb Currencies Plastic Card", "Network not stated in official text", "Prepaid multi-currency", "Current product confirmed", "SAR 75; current campaign page also advertises lifetime annual-fee waiver without permanence criteria", "No loyalty rate published", "Conflict pending", "Standard tariff fee versus promotional lifetime-fee waiver; network not stated", "Preserved standard and promotional evidence", "https://anb.com.sa/web/anb/currencies-card", CURRENCIES_TC, CURRENCIES_TC),
    row("ANB-17", "anb Currencies Digital Card", "Network not stated in official text", "Digital prepaid multi-currency", "Current product confirmed", "SAR 50", "No loyalty rate published", "Missing evidence", "Network and explicit non-participation in rewards are not stated", "No inferred network/reward value", "https://anb.com.sa/web/anb/currencies-card", CURRENCIES_TC, CURRENCIES_TC),
    row("ANB-18", "anb Mada Infinite Card", "mada; co-badge not stated", "Debit", "Current tier confirmed", "No tier-specific annual fee published", "250 points for qualifying mada e-commerce transaction; unit wording incomplete", "Missing evidence", "Tier-specific pricing, co-badge network and unambiguous earning denominator missing", "No inferred values", "https://anb.com.sa/web/anb/mada-cards", MADA_TC, REWARDS),
    row("ANB-19", "anb Mada Platinum Card", "mada", "Debit", "Current tier confirmed", "No tier-specific annual fee published", "250 points for qualifying mada e-commerce transaction; unit wording incomplete", "Missing evidence", "Tier-specific pricing and unambiguous earning denominator missing", "No inferred values", "https://anb.com.sa/web/anb/mada-cards", MADA_TC, REWARDS),
    row("ANB-20", "anb Mada Gold Card", "mada", "Debit", "Current tier confirmed", "No tier-specific annual fee published", "250 points for qualifying mada e-commerce transaction; unit wording incomplete", "Missing evidence", "Tier-specific pricing and unambiguous earning denominator missing", "No inferred values", "https://anb.com.sa/web/anb/mada-cards", MADA_TC, REWARDS),
    row("ANB-21", "anb Mada Classic Card", "mada", "Debit", "Current tier confirmed", "No tier-specific annual fee published", "250 points for qualifying mada e-commerce transaction; unit wording incomplete", "Missing evidence", "Tier-specific pricing and unambiguous earning denominator missing", "No inferred values", "https://anb.com.sa/web/anb/mada-cards", MADA_TC, REWARDS),
    row("ANB-22", "anb Corporate Card", "Visa", "Corporate credit", "Current product confirmed", "No corporate pricing schedule located", "No corporate reward schedule located", "Missing evidence", "Current corporate pricing/fees and rewards applicability document missing", "No inferred fee/reward value", "https://anb.com.sa/en/web/anb/corporate-credit-card", CORP_TC, CORP_TC),
    row("ANB-23", "anb Business Card", "Visa", "Business credit", "Current product confirmed", "No business pricing schedule located", "No business reward schedule located", "Missing evidence", "Current Business Card pricing/fees and rewards applicability document missing", "No inferred fee/reward value", "https://anb.com.sa/en/web/anb/corporate-credit-card", CORP_TC, CORP_TC),
    row("ANB-24", "anb Purchase Card", "Visa", "Purchasing credit", "Current product confirmed", "No purchase-card pricing schedule located", "No purchase-card reward schedule located", "Missing evidence", "Current Purchase Card pricing/fees and rewards applicability document missing", "No inferred fee/reward value", "https://anb.com.sa/en/web/anb/corporate-credit-card", CORP_TC, CORP_TC),
]


def main() -> None:
    wb = load_workbook(WORKBOOK)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.append(HEADERS)
    for item in ROWS:
        ws.append(item)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index in (11, 12, 13, 14):
            cell = row_cells[index]
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    widths = [12, 38, 18, 22, 30, 42, 40, 24, 55, 35, 15, 45, 45, 45, 45]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    ws.row_dimensions[1].height = 42
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 72

    table = Table(displayName="ANBValidation20260730", ref=f"A1:O{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)

    temp = WORKBOOK.with_suffix(".tmp.xlsx")
    wb.save(temp)
    temp.replace(WORKBOOK)
    print(f"Added {len(ROWS)} ANB validation rows to {WORKBOOK}")


if __name__ == "__main__":
    main()
